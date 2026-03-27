"""
output/excel_writer.py
Writes extracted records to a consolidated Excel file matching the
sample output.xlsx format (13 columns).

Columns: Sr.no., Village, Survey No., Area in NA Order, Dated,
         NA Order No., Lease Deed Doc. No., Lease Area, Lease Start,
         Tenure, Validity (till), e-Challan No., Valid Up to

Records from NA Permission and eChallan documents are grouped by
Survey Number into single consolidated rows.
"""

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from config import CONSOLIDATED_COLUMNS

logger = logging.getLogger(__name__)


def _normalize_survey_no(val: str) -> str:
    """Normalize survey number for matching: lowercase, strip spaces."""
    if not val:
        return ""
    return re.sub(r"\s+", "", str(val)).lower().strip()


def _extract_survey_from_filename(filename: str) -> str:
    """
    Try to extract survey number from the filename.
    Examples:
        "251-p2 FINAL ORDER.pdf" → "251/p2"
        "Rampura Mota S.No.- 251p2 Lease Deed No.- 141.pdf" → "251/p2"
        "257 FINAL ORDER.pdf" → "257"
    """
    name = Path(filename).stem.lower()

    # Pattern: "s.no.- 251p2" or "s.no.-255"
    match = re.search(r's\.?no\.?\s*-?\s*(\d+)\s*/?p?(\d*)', name)
    if match:
        num = match.group(1)
        part = match.group(2)
        return f"{num}/p{part}" if part else num

    # Pattern: "251-p2" at start
    match = re.search(r'^(\d+)\s*-?\s*p(\d+)', name)
    if match:
        return f"{match.group(1)}/p{match.group(2)}"

    # Pattern: just a number at start like "257"
    match = re.search(r'^(\d+)\s', name)
    if match:
        return match.group(1)

    return ""


class ExcelWriter:
    """
    Collects extracted records and writes a consolidated Excel report.

    Usage:
        writer = ExcelWriter()
        writer.add(record)  # add each extracted record
        writer.save("output.xlsx")
    """

    def __init__(self):
        self._na_records: List[Dict] = []
        self._echallan_records: List[Dict] = []
        self._errors: int = 0

    def add(self, record: Dict) -> None:
        """Add an extracted record to the writer."""
        if record.get("_error"):
            self._errors += 1

        doc_type = record.get("doc_type", "unknown")
        if doc_type == "na_permission":
            self._na_records.append(record)
        elif doc_type == "echallan":
            self._echallan_records.append(record)
        else:
            logger.warning("Unknown doc_type '%s' for %s", doc_type, record.get("source_file"))
            self._echallan_records.append(record)

    def save(self, filepath: str) -> None:
        """
        Write consolidated Excel file.

        Groups NA + eChallan records by Survey Number and writes
        one row per survey to match the sample output.xlsx format.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated Report"

        # ── Write header row ───────────────────────────────────────────────────
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(CONSOLIDATED_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # ── Build consolidated rows ────────────────────────────────────────────
        consolidated = self._build_consolidated_rows()

        for row_idx, row_data in enumerate(consolidated, start=2):
            for col_idx, col_name in enumerate(CONSOLIDATED_COLUMNS, start=1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # ── Auto-size columns ──────────────────────────────────────────────────
        for col_idx in range(1, len(CONSOLIDATED_COLUMNS) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

        # Save
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        logger.info("Excel report saved to %s (%d rows)", filepath, len(consolidated))

    def _build_consolidated_rows(self) -> List[Dict]:
        """
        Group NA and eChallan records by Survey Number.

        For each survey number, combine:
          - NA fields: Area in NA Order, Dated, NA Order No.
          - eChallan/LeaseDeed fields: Lease Deed Doc. No., Lease Area, Lease Start,
            Tenure, Validity (till), e-Challan No., Valid Up to
        """
        # Index NA records by normalized survey number
        na_by_survey = {}
        for rec in self._na_records:
            survey = rec.get("survey_number", "") or _extract_survey_from_filename(rec.get("source_file", ""))
            key = _normalize_survey_no(survey)
            if key:
                na_by_survey[key] = rec

        # Index eChallan records by normalized survey number
        echallan_by_survey = {}
        for rec in self._echallan_records:
            survey = rec.get("survey_number", "") or _extract_survey_from_filename(rec.get("source_file", ""))
            key = _normalize_survey_no(survey)
            if key:
                echallan_by_survey[key] = rec

        # Merge all unique survey numbers
        all_surveys = set(list(na_by_survey.keys()) + list(echallan_by_survey.keys()))
        rows = []

        for idx, survey_key in enumerate(sorted(all_surveys), start=1):
            na = na_by_survey.get(survey_key, {})
            ec = echallan_by_survey.get(survey_key, {})

            # Determine display survey number (prefer NA record's value)
            survey_display = na.get("survey_number") or ec.get("survey_number") or survey_key

            # Tenure can come from NA (lease_term) or eChallan (tenure_years)
            tenure = ec.get("tenure_years", "") or na.get("lease_term", "")

            row = {
                "Sr.no.": idx,
                "Village": na.get("village", "") or ec.get("village", "") or "Rampura Mota",
                "Survey No.": survey_display,
                "Area in NA Order": na.get("land_area", ""),
                "Dated": na.get("order_date", ""),
                "NA Order No.": na.get("order_number", ""),
                "Lease Deed Doc. No.": ec.get("lease_deed_doc_no", "") or ec.get("challan_number", ""),
                "Lease Area": ec.get("lease_area", "") or ec.get("land_area", ""),
                "Lease Start": ec.get("lease_start_date", "") or ec.get("violation_date", ""),
                "Tenure": tenure,
                "Validity (till)": ec.get("validity_till", "") or ec.get("valid_up_to", ""),
                "e-Challan No.": ec.get("echallan_number", ""),
                "Valid Up to": ec.get("valid_up_to", ""),
            }
            rows.append(row)

        return rows

    def summary(self) -> Dict:
        """Return summary counts for the CLI."""
        return {
            "na_permissions": len(self._na_records),
            "echallans": len(self._echallan_records),
            "errors": self._errors,
            "total": len(self._na_records) + len(self._echallan_records),
        }